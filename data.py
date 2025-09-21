import sqlite3
from datetime import datetime
import openpyxl

# === Koneksi Database ===
conn = sqlite3.connect("koperasi.db")
cur = conn.cursor()

# === Buat Tabel ===
cur.execute("""
CREATE TABLE IF NOT EXISTS anggota (
    id_anggota INTEGER PRIMARY KEY AUTOINCREMENT,
    nama TEXT NOT NULL,
    alamat TEXT,
    no_hp TEXT,
    tanggal_gabung TEXT
)
""")

cur.execute("""
CREATE TABLE IF NOT EXISTS pinjaman (
    id_pinjaman INTEGER PRIMARY KEY AUTOINCREMENT,
    id_anggota INTEGER,
    pinjaman_pokok REAL,
    total_pinjaman REAL,
    sisa_pinjaman REAL,
    tempo_hari INTEGER,
    status TEXT,
    FOREIGN KEY(id_anggota) REFERENCES anggota(id_anggota)
)
""")

cur.execute("""
CREATE TABLE IF NOT EXISTS angsuran (
    id_angsuran INTEGER PRIMARY KEY AUTOINCREMENT,
    tanggal TEXT,
    id_pinjaman INTEGER,
    id_anggota INTEGER,
    jumlah_bayar REAL,
    sisa_pinjaman REAL,
    status TEXT,
    FOREIGN KEY(id_pinjaman) REFERENCES pinjaman(id_pinjaman),
    FOREIGN KEY(id_anggota) REFERENCES anggota(id_anggota)
)
""")

conn.commit()

# === Fungsi CRUD ===
def tambah_anggota(nama, alamat, no_hp):
    tanggal = datetime.now().strftime("%Y-%m-%d")
    cur.execute("INSERT INTO anggota (nama, alamat, no_hp, tanggal_gabung) VALUES (?, ?, ?, ?)",
                (nama, alamat, no_hp, tanggal))
    conn.commit()
    print(f"[✓] Anggota '{nama}' berhasil ditambahkan.\n")

def tambah_pinjaman(id_anggota, jumlah, tempo):
    # Hitung pinjaman dengan bunga 20%
    total_pinjaman = jumlah * 1.2  

    cur.execute("INSERT INTO pinjaman (id_anggota, pinjaman_pokok, total_pinjaman, sisa_pinjaman, tempo_hari, status) VALUES (?, ?, ?, ?, ?, ?)",
                (id_anggota, jumlah, total_pinjaman, total_pinjaman, tempo, "Belum Lunas"))
    conn.commit()
    print(f"[✓] Pinjaman Rp {jumlah:,.0f} + bunga 20% = Rp {total_pinjaman:,.0f}, Tempo {tempo} hari. Untuk anggota ID {id_anggota} berhasil ditambahkan.\n")

def catat_angsuran(id_pinjaman, jumlah_bayar):
    tanggal = datetime.now().strftime("%Y-%m-%d")

    cur.execute("SELECT id_anggota, sisa_pinjaman, tempo_hari FROM pinjaman WHERE id_pinjaman = ?", (id_pinjaman,))
    data = cur.fetchone()
    if not data:
        print("[!] Pinjaman tidak ditemukan!\n")
        return
    id_anggota, sisa, tempo = data

    # Hitung sisa pinjaman
    sisa_baru = sisa - jumlah_bayar
    status = "Lunas" if sisa_baru <= 0 else "Belum Lunas"

    # Kurangi tempo setiap kali bayar
    tempo_baru = max(tempo - 1, 0)
    if tempo_baru == 0 and sisa_baru > 0:
        status = "Jatuh Tempo"

    # Update pinjaman
    cur.execute("UPDATE pinjaman SET sisa_pinjaman = ?, tempo_hari = ?, status = ? WHERE id_pinjaman = ?",
                (max(sisa_baru, 0), tempo_baru, status, id_pinjaman))

    # Catat angsuran
    cur.execute("INSERT INTO angsuran (tanggal, id_pinjaman, id_anggota, jumlah_bayar, sisa_pinjaman, status) VALUES (?, ?, ?, ?, ?, ?)",
                (tanggal, id_pinjaman, id_anggota, jumlah_bayar, max(sisa_baru, 0), status))
    conn.commit()
    print(f"[✓] Angsuran Rp {jumlah_bayar:,.0f} dicatat. Sisa: Rp {max(sisa_baru,0):,.0f}, Tempo sisa: {tempo_baru} hari, Status: {status}\n")

def lihat_anggota():
    cur.execute("SELECT * FROM anggota")
    data = cur.fetchall()
    print("=== Data Anggota ===")
    for row in data:
        print(row)
    print()

def lihat_pinjaman():
    cur.execute("SELECT * FROM pinjaman")
    data = cur.fetchall()
    print("=== Data Pinjaman ===")
    for row in data:
        print(row)
        # 🔔 Tambahkan peringatan otomatis
        id_pinjaman, id_anggota, pokok, total, sisa, tempo, status = row
        if status == "Belum Lunas" and tempo <= 3 and tempo > 0:
            print(f"   ⚠️ Peringatan: Pinjaman ID {id_pinjaman} (Anggota {id_anggota}) tinggal {tempo} hari lagi!")
        elif status == "Jatuh Tempo":
            print(f"   ❌ Pinjaman ID {id_pinjaman} sudah jatuh tempo!")
    print()

def lihat_angsuran():
    cur.execute("SELECT * FROM angsuran")
    data = cur.fetchall()
    print("=== Data Angsuran ===")
    for row in data:
        print(row)
    print()

# === Fitur Rekap Bulanan ===
def rekap_bulanan(bulan_tahun):
    print(f"\n=== Rekap Bulan {bulan_tahun} ===")

    cur.execute("SELECT SUM(total_pinjaman) FROM pinjaman WHERE substr(rowid,1,7) = ?", (bulan_tahun,))
    total_pinjaman = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(jumlah_bayar) FROM angsuran WHERE substr(tanggal,1,7) = ?", (bulan_tahun,))
    total_angsuran = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(sisa_pinjaman) FROM pinjaman")
    total_sisa = cur.fetchone()[0] or 0

    print(f"Total Pinjaman    : Rp {total_pinjaman:,.0f}")
    print(f"Total Angsuran    : Rp {total_angsuran:,.0f}")
    print(f"Total Sisa Hutang : Rp {total_sisa:,.0f}\n")

# === Daftar Jatuh Tempo + Ekspor Excel ===
def daftar_jatuh_tempo():
    cur.execute("SELECT * FROM pinjaman WHERE status = 'Jatuh Tempo'")
    data = cur.fetchall()
    print("=== Daftar Pinjaman Jatuh Tempo ===")
    if not data:
        print("Tidak ada pinjaman jatuh tempo.\n")
    else:
        for row in data:
            print(row)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jatuh Tempo"

        headers = ["ID Pinjaman", "ID Anggota", "Pinjaman Pokok", "Total Pinjaman", 
                   "Sisa Pinjaman", "Tempo (hari)", "Status"]
        ws.append(headers)

        for row in data:
            ws.append(row)

        filename = f"jatuh_tempo_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)
        print(f"[✓] Data jatuh tempo berhasil diekspor ke file: {filename}\n")

# === Ekspor Rekap Bulanan ke Excel ===
def ekspor_rekap_bulanan(bulan_tahun):
    cur.execute("SELECT SUM(total_pinjaman) FROM pinjaman WHERE substr(rowid,1,7) = ?", (bulan_tahun,))
    total_pinjaman = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(jumlah_bayar) FROM angsuran WHERE substr(tanggal,1,7) = ?", (bulan_tahun,))
    total_angsuran = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(sisa_pinjaman) FROM pinjaman")
    total_sisa = cur.fetchone()[0] or 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Bulanan"

    headers = ["Bulan", "Total Pinjaman", "Total Angsuran", "Total Sisa Hutang"]
    ws.append(headers)
    ws.append([bulan_tahun, total_pinjaman, total_angsuran, total_sisa])

    filename = f"rekap_bulanan_{bulan_tahun}.xlsx"
    wb.save(filename)
    print(f"[✓] Rekap Bulanan {bulan_tahun} berhasil diekspor ke file: {filename}\n")

# === Ekspor Rekap Per Anggota ke Excel ===
def ekspor_rekap_anggota(id_anggota):
    # Ambil data anggota
    cur.execute("SELECT * FROM anggota WHERE id_anggota = ?", (id_anggota,))
    anggota = cur.fetchone()
    if not anggota:
        print("[!] Anggota tidak ditemukan!\n")
        return

    # Ambil semua pinjaman anggota
    cur.execute("SELECT * FROM pinjaman WHERE id_anggota = ?", (id_anggota,))
    pinjaman = cur.fetchall()

    # Ambil semua angsuran anggota
    cur.execute("SELECT * FROM angsuran WHERE id_anggota = ?", (id_anggota,))
    angsuran = cur.fetchall()

    # Buat file Excel
    wb = openpyxl.Workbook()

    # Sheet 1: Data Anggota
    ws1 = wb.active
    ws1.title = "Data Anggota"
    ws1.append(["ID Anggota", "Nama", "Alamat", "No HP", "Tanggal Gabung"])
    ws1.append(anggota)

    # Sheet 2: Pinjaman
    ws2 = wb.create_sheet("Pinjaman")
    ws2.append(["ID Pinjaman", "ID Anggota", "Pinjaman Pokok", "Total Pinjaman", 
                "Sisa Pinjaman", "Tempo (hari)", "Status"])
    for row in pinjaman:
        ws2.append(row)

    # Sheet 3: Angsuran
    ws3 = wb.create_sheet("Angsuran")
    ws3.append(["ID Angsuran", "Tanggal", "ID Pinjaman", "ID Anggota", 
                "Jumlah Bayar", "Sisa Pinjaman", "Status"])
    for row in angsuran:
        ws3.append(row)

    filename = f"rekap_anggota_{id_anggota}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"[✓] Rekap anggota ID {id_anggota} berhasil diekspor ke file: {filename}\n")

# === Menu Interaktif ===
def menu():
    while True:
        print("""
=== MENU KOPERASI HARIAN ===
1. Tambah Anggota
2. Tambah Pinjaman
3. Catat Angsuran Harian
4. Lihat Data Anggota
5. Lihat Data Pinjaman
6. Lihat Data Angsuran
7. Rekap Bulanan
8. Daftar Pinjaman Jatuh Tempo (Excel)
9. Ekspor Rekap Bulanan ke Excel
10. Ekspor Rekap Per Anggota ke Excel
0. Keluar
""")
        pilih = input("Pilih menu: ")

        if pilih == "1":
            nama = input("Nama: ")
            alamat = input("Alamat: ")
            no_hp = input("No HP: ")
            tambah_anggota(nama, alamat, no_hp)

        elif pilih == "2":
            id_anggota = int(input("ID Anggota: "))
            jumlah = float(input("Jumlah Pinjaman: "))
            print("Pilih Tempo: 24 / 30 hari")
            tempo = int(input("Tempo (hari): "))
            if tempo not in [24, 30]:
                print("[!] Tempo hanya boleh 24 atau 30 hari.\n")
            else:
                tambah_pinjaman(id_anggota, jumlah, tempo)

        elif pilih == "3":
            id_pinjaman = int(input("ID Pinjaman: "))
            bayar = float(input("Jumlah Bayar: "))
            catat_angsuran(id_pinjaman, bayar)

        elif pilih == "4":
            lihat_anggota()

        elif pilih == "5":
            lihat_pinjaman()

        elif pilih == "6":
            lihat_angsuran()

        elif pilih == "7":
            bulan_tahun = input("Masukkan bulan (format YYYY-MM): ")
            rekap_bulanan(bulan_tahun)

        elif pilih == "8":
            daftar_jatuh_tempo()

        elif pilih == "9":
            bulan_tahun = input("Masukkan bulan (format YYYY-MM): ")
            ekspor_rekap_bulanan(bulan_tahun)

        elif pilih == "10":
            id_anggota = int(input("Masukkan ID Anggota: "))
            ekspor_rekap_anggota(id_anggota)

        elif pilih == "0":
            print("Keluar dari program...")
            break

        else:
            print("[!] Pilihan tidak valid.\n")

if __name__ == "__main__":
    menu()
